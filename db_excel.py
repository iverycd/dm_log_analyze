from openpyxl import Workbook
from openpyxl.styles import PatternFill, colors, Font, Alignment


def read_db_to_xlsx(sql_result_list, sql_result_list2, excelname):
    # 循环数据写入工作表1内容
    jb_date_lists = sql_result_list
    jb_date_list = jb_date_lists[0]
    descripte = jb_date_lists[1]
    excel_rows_num = jb_date_lists[2]
    # 写入工作表2内容
    jb_date_lists2 = sql_result_list2
    jb_date_list2 = jb_date_lists2[0]
    descripte2 = jb_date_lists2[1]
    excel_rows_num2 = jb_date_lists2[2]
    # 要创建的xlsx名称
    dest_filename = excelname
    # 新建工作簿
    wb = Workbook()
    # 新建工作表1
    ws1 = wb.active
    ws1.title = "最大执行耗时排序"
    # 工作表1列名
    for i in range(0, len(descripte)):
        ws1.cell(row=1, column=i + 1, value=descripte[i])
    # 工作表1写入数据
    for i in range(2, len(jb_date_list) + 2):
        for j in range(0, len(descripte)):
            if jb_date_list[i - 2][j] is None:
                ws1.cell(row=i, column=j + 1, value='')
            else:
                ws1.cell(row=i, column=j + 1, value=jb_date_list[i - 2][j])
    # 工作表1设置单元格背景
    fill_a1 = PatternFill("solid", fgColor="1874CD")
    ws1["A1"].fill = fill_a1
    fill_b1 = PatternFill("solid", fgColor="1874CD")
    ws1["B1"].fill = fill_b1
    fill_c1 = PatternFill("solid", fgColor="1874CD")
    ws1["C1"].fill = fill_c1
    fill_d1 = PatternFill("solid", fgColor="1874CD")
    ws1["D1"].fill = fill_d1
    fill_e1 = PatternFill("solid", fgColor="1874CD")
    ws1["E1"].fill = fill_e1
    fill_f1 = PatternFill("solid", fgColor="1874CD")
    ws1["F1"].fill = fill_f1
    fill_g1 = PatternFill("solid", fgColor="1874CD")
    ws1["G1"].fill = fill_g1
    fill_h1 = PatternFill("solid", fgColor="1874CD")
    ws1["H1"].fill = fill_h1
    fill_i1 = PatternFill("solid", fgColor="1874CD")
    ws1["I1"].fill = fill_i1
    fill_j1 = PatternFill("solid", fgColor="1874CD")
    ws1["J1"].fill = fill_j1
    # 工作表1设置单元格文字颜色
    font_set = Font(color=colors.WHITE, bold=True)
    ws1['A1'].font = font_set
    ws1['B1'].font = font_set
    ws1['C1'].font = font_set
    ws1['D1'].font = font_set
    ws1['E1'].font = font_set
    ws1['F1'].font = font_set
    ws1['G1'].font = font_set
    ws1['H1'].font = font_set
    ws1['I1'].font = font_set
    ws1['J1'].font = font_set
    # 工作表1设置列宽
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 100
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 20
    ws1.column_dimensions['E'].width = 25
    ws1.column_dimensions['F'].width = 20
    ws1.column_dimensions['G'].width = 10
    ws1.column_dimensions['H'].width = 10
    ws1.column_dimensions['I'].width = 100
    ws1.column_dimensions['J'].width = 20
    # 工作表1设置单元格对齐,替换参数的列为填充fill避免文本太长全部显示出来
    ws_area = ws1["I2:I%s" % excel_rows_num]
    for i in ws_area:
        for j in i:
            j.alignment = Alignment(horizontal='fill', vertical='center', wrap_text=False)
    # 工作表1其余列为左对齐
    ws_area = ws1["A2:H%s" % excel_rows_num]
    for i in ws_area:
        for j in i:
            j.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    # 工作表1冻结首行，方便下拉的时候能一直显示列名,设置A1没效果，所以要设置为A2即A2之上的第一行冻结
    ws1.freeze_panes = 'A2'
    # 以上是工作表1完毕
    # 以下是新建工作表2
    ws2 = wb.create_sheet("最多执行次数排序")
    # 工作表2列名
    for i in range(0, len(descripte2)):
        ws2.cell(row=1, column=i + 1, value=descripte2[i])
    # 工作表2写入数据
    for i in range(2, len(jb_date_list2) + 2):
        for j in range(0, len(descripte2)):
            if jb_date_list2[i - 2][j] is None:
                ws2.cell(row=i, column=j + 1, value='')
            else:
                ws2.cell(row=i, column=j + 1, value=jb_date_list2[i - 2][j])
    # 设置工作表2的背景
    ws2["A1"].fill = fill_a1
    ws2["B1"].fill = fill_b1
    ws2["C1"].fill = fill_c1
    ws2["D1"].fill = fill_d1
    ws2["E1"].fill = fill_e1
    ws2["F1"].fill = fill_f1
    ws2["G1"].fill = fill_g1
    ws2["H1"].fill = fill_h1
    ws2["I1"].fill = fill_i1
    ws2["J1"].fill = fill_j1
    # 设置工作表2的文字颜色
    ws2['A1'].font = font_set
    ws2['B1'].font = font_set
    ws2['C1'].font = font_set
    ws2['D1'].font = font_set
    ws2['E1'].font = font_set
    ws2['F1'].font = font_set
    ws2['G1'].font = font_set
    ws2['H1'].font = font_set
    ws2['I1'].font = font_set
    ws2['J1'].font = font_set
    # 设置工作表2的列宽
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 100
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 20
    ws2.column_dimensions['E'].width = 25
    ws2.column_dimensions['F'].width = 20
    ws2.column_dimensions['G'].width = 10
    ws2.column_dimensions['H'].width = 10
    ws2.column_dimensions['I'].width = 100
    ws2.column_dimensions['J'].width = 20
    # 工作表2设置单元格对齐,替换参数的列为填充fill避免文本太长全部显示出来
    ws_area = ws2["I2:I%s" % excel_rows_num2]
    for i in ws_area:
        for j in i:
            j.alignment = Alignment(horizontal='fill', vertical='center', wrap_text=False)
    # 工作表2,其余列为左对齐
    ws_area = ws2["A2:H%s" % excel_rows_num2]
    for i in ws_area:
        for j in i:
            j.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
    # 工作表2,冻结首行，方便下拉的时候能一直显示列名,设置A1没效果，所以要设置为A2即A2之上的第一行冻结
    ws2.freeze_panes = 'A2'
    # 创建xlsx
    wb.save(filename=dest_filename)
